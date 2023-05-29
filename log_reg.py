import pandas as pd
from openpyxl import load_workbook
import nltk
from nltk.tokenize import sent_tokenize
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import pymorphy2
from sklearn.feature_extraction.text import CountVectorizer
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report

nb = GaussianNB()
nltk.download('stopwords')
nltk.download('punkt')
unique_stops = set(stopwords.words('russian'))
morph = pymorphy2.MorphAnalyzer()
count = CountVectorizer()


workbook_with_poems = load_workbook('Выборка.xlsx')

akmeists_sheet = workbook_with_poems['Акмеисты']
symbolists_sheet = workbook_with_poems['Символисты']
futurists_sheet = workbook_with_poems['Футуристы']

def lemmatize(text):
    list_of_words = word_tokenize(text)
    list_of_lemmas = []
    for i in range(len(list_of_words)):
        lemma = morph.parse(list_of_words[i])[0].normal_form
        list_of_lemmas.append(lemma)
    text_lemmatized = ' '.join(list_of_lemmas)
    return text_lemmatized


def remove_stopwords(text):
    text_tokens = word_tokenize(text)
    remove_sw = [word for word in text_tokens if word not in unique_stops]
    return remove_sw


def preprocessing(text):
    text = text.lower()
    text = text.strip()
    text = lemmatize(text)
    text = remove_stopwords(text)
    while '.' in text:
        text.remove('.')
    return text

poetry_dicts = []
texts_f = []

for i in range(2,101):
    text = futurists_sheet.cell(i, column=5).value
    author = futurists_sheet.cell(i, column=1).value
    year_of_creation = futurists_sheet.cell(i, column=4).value
    if author != 'В.В. Хлебников':
        poem_dict = {'text': text, 'style': 'Футуризм'}
        poetry_dicts.append(poem_dict)
        texts_f.append(text)

texts_s = []
for i in range(2,132):
    text = symbolists_sheet.cell(i, column=5).value
    author = symbolists_sheet.cell(i, column=1).value
    year_of_creation = symbolists_sheet.cell(i, column=4).value
    poem_dict = {'text': text, 'style': 'Символизм'}
    poetry_dicts.append(poem_dict)
    texts_s.append(text)

texts_a = []
for i in range(2,102):
    text = akmeists_sheet.cell(i, column=5).value
    author = akmeists_sheet.cell(i, column=1).value
    year_of_creation = akmeists_sheet.cell(i, column=4).value
    poem_dict = {'text': text, 'style': 'Акмеизм'}
    poetry_dicts.append(poem_dict)
    texts_a.append(text)

poem_df = pd.DataFrame(poetry_dicts)

print(poem_df)

text_proc = []

for poem in poem_df['text']:
    poem_proc = preprocessing(poem)
    poem_proc =' '.join(poem_proc)
    text_proc.append(poem_proc)

matrix = count.fit_transform(text_proc).toarray()

X = matrix

Y = poem_df['style'].values

x_train, x_test, y_train, y_test = train_test_split(X, Y, test_size = 0.33, random_state = 42)

result_bayes = nb.fit(x_train, y_train)

Y_pred = nb.predict(x_test)


print(classification_report(y_test,Y_pred))
print(confusion_matrix(y_test,Y_pred))