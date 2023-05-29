from navec import Navec
from scipy import spatial
import numpy as np
import re
import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import pymorphy2
from collections import defaultdict
from openpyxl import load_workbook
import zipfile
import wget
import gensim
import sys
import os
from ufal.udpipe import Model, Pipeline
import yake
import string
from openpyxl import load_workbook

nltk.download('stopwords')
nltk.download('punkt')
unique_stops = set(stopwords.words('russian'))
morph = pymorphy2.MorphAnalyzer()
udpipe_url = 'https://rusvectores.org/static/models/udpipe_syntagrus.model'
modelfile = wget.download(udpipe_url)
punctuation = string.punctuation
language = "ru"
max_ngram_size = 1
deduplication_threshold = 0.9
numOfKeywords = 8
kw_extractor = yake.KeywordExtractor(lan=language, n=max_ngram_size, dedupLim=deduplication_threshold, top=numOfKeywords, features=None)

model_url = 'http://vectors.nlpl.eu/repository/20/180.zip'

model_file = model_url.split('/')[-1]
with zipfile.ZipFile(model_file, 'r') as archive:
   stream = archive.open('model.bin')
   model = gensim.models.KeyedVectors.load_word2vec_format(stream, binary=True)

def lemmatize(text):
    list_of_words = word_tokenize(text)
    list_of_lemmas = []
    for i in range(len(list_of_words)):
        lemma = morph.parse(list_of_words[i])[0].normal_form
        list_of_lemmas.append(lemma)
    text_lemmatized = ' '.join(list_of_lemmas)
    return text_lemmatized

def remove_stopwords(text):
    text_tokens = word_tokenize(text)
    remove_sw = [word for word in text_tokens if word not in unique_stops]
    return remove_sw

def preprocessing(text):
    text = text.lower()
    text = text.strip()
    text = lemmatize(text)
    text = remove_stopwords(text)
    while '.' in text:
        text.remove('.')
    return text

def process(pipeline, text, keep_pos=True, keep_punct=False):
   entities = {'PROPN'}
   named = False  # переменная для запоминания того, что нам встретилось имя собственное
   memory = []
   mem_case = None
   mem_number = None
   tagged_propn = []
   # обрабатываем текст, получаем результат в формате conllu:
   processed = pipeline.process(text)
   # пропускаем строки со служебной информацией:
   content = [l for l in processed.split('\n') if not l.startswith('#')]
   # извлекаем из обработанного текста леммы, тэги и морфологические характеристики
   tagged = [w.split('\t') for w in content if w]
   for t in tagged:
       if len(t) != 10: # если список короткий — строчка не содержит разбора, пропускаем
           continue
       (word_id,token,lemma,pos,xpos,feats,head,deprel,deps,misc) = t
       if not lemma or not token: # если слово пустое — пропускаем
           continue
       if pos in entities: # здесь отдельно обрабатываем имена собственные — они требуют особого обращения
           if '|' not in feats:
               tagged_propn.append('%s_%s' % (lemma, pos))
               continue
           morph = {el.split('=')[0]: el.split('=')[1] for el in feats.split('|')}
           if 'Case' not in morph or 'Number' not in morph:
               tagged_propn.append('%s_%s' % (lemma, pos))
               continue
           if not named:
               named = True
               mem_case = morph['Case']
               mem_number = morph['Number']
           if morph['Case'] == mem_case and morph['Number'] == mem_number:
               memory.append(lemma)
               if 'SpacesAfter=\\n' in misc or 'SpacesAfter=\s\\n' in misc:
                   named = False
                   past_lemma = '::'.join(memory)
                   memory = []
                   tagged_propn.append(past_lemma + '_PROPN ')
           else:
               named = False
               past_lemma = '::'.join(memory)
               memory = []
               tagged_propn.append(past_lemma + '_PROPN ')
               tagged_propn.append('%s_%s' % (lemma, pos))
       else:
           if not named:
               if pos == 'NUM' and token.isdigit():  # Заменяем числа на xxxxx той же длины
                   lemma = num_replace(token)
               tagged_propn.append('%s_%s' % (lemma, pos))
           else:
               named = False
               past_lemma = '::'.join(memory)
               memory = []
               tagged_propn.append(past_lemma + '_PROPN ')
               tagged_propn.append('%s_%s' % (lemma, pos))

   if not keep_punct: # обрабатываем случай, когда пользователь попросил не сохранять пунктуацию (по умолчанию она сохраняется)
       tagged_propn = [word for word in tagged_propn if word.split('_')[1] != 'PUNCT']
   if not keep_pos:
       tagged_propn = [word.split('_')[0] for word in tagged_propn]
   return tagged_propn

def tag_ud(text, modelfile='udpipe_syntagrus.model'):
   udpipe_model_url = 'https://rusvectores.org/static/models/udpipe_syntagrus.model'
   udpipe_filename = udpipe_model_url.split('/')[-1]
   if not os.path.isfile(modelfile):
       print('UDPipe model not found. Downloading...', file=sys.stderr)
       wget.download(udpipe_model_url)
   print('\nLoading the model...', file=sys.stderr)
   model = Model.load(modelfile)
   process_pipeline = Pipeline(model, 'tokenize', Pipeline.DEFAULT, Pipeline.DEFAULT, 'conllu')
   print('Processing input...', file=sys.stderr)
   lines = text.split('\n')
   tagged = []
   for line in lines:
       # line = unify_sym(line.strip()) # здесь могла бы быть ваша функция очистки текста
       output = process(process_pipeline, text=line)
       tagged_line = ' '.join(output)
       tagged.append(tagged_line)
   return '\n'.join(tagged)

def get_keywords(text):
    proc_text = ' '.join(preprocessing(text))
    keywords = kw_extractor.extract_keywords(proc_text)
    keywords_list = [pair[0] for pair in keywords]
    return keywords_list

f1 = open('Futurists processed for RusVectores.txt', encoding='utf-8')

topics_futurism = f1.readlines()

topics_futurism = [topic.split() for topic in topics_futurism]

f2 = open('Symbolists processed for RusVectores.txt', encoding='utf-8')

topics_symbolism = f2.readlines()

topics_symbolism = [topic.split() for topic in topics_symbolism]

f3 = open('Akmeists processed for RusVectores.txt', encoding='utf-8')

topics_akmeism = f3.readlines()

topics_akmeism = [topic.split() for topic in topics_akmeism]


def topic_similarity(text,topic):
    keywords = get_keywords(text)
    keywords = [tag_ud(word) for word in keywords]
    similarity_scores = []
    missed = 0
    for word in topic:
        for keyword in keywords:
            try:
                score = model.similarity(word, keyword)
                similarity_scores.append(score)
            except:
                missed +=1
    print('Пропущено', missed)
    avg_score = sum(similarity_scores)/len(similarity_scores)
    return avg_score

def many_topics_similarity(text,list_of_topics):
    scores = []
    for topic in list_of_topics:
        scores.append(topic_similarity(text,topic))
    avg_score = sum(scores)/len(scores)
    return avg_score

def get_style(text):
    futurism_score = many_topics_similarity(text, topics_futurism)
    symbolism_score = many_topics_similarity(text, topics_symbolism)
    akmeism_score = many_topics_similarity(text, topics_akmeism)
    scores = [futurism_score, symbolism_score, akmeism_score]
    dominant = max(scores)
    if dominant == futurism_score:
        print('Футуризм')
    elif dominant == symbolism_score:
        print('Символизм')
    elif dominant == akmeism_score:
        print("Акмеизм")

def get_style_from_scores(scores: dict):
    futurism_score = scores['Футуризм']
    symbolism_score = scores['Символизм']
    akmeism_score = scores['Акмеизм']
    scores = [futurism_score, symbolism_score, akmeism_score]
    dominant = max(scores)
    if dominant == futurism_score:
        return 'Футуризм'
    elif dominant == symbolism_score:
        return 'Символизм'
    elif dominant == akmeism_score:
        return "Акмеизм"

poetry_dicts = []

workbook_with_poems = load_workbook('Выборка.xlsx')
futurists_sheet = workbook_with_poems['Футуристы']
texts_f = []
for i in range(2,101):
    text = futurists_sheet.cell(i, column=5).value
    author = futurists_sheet.cell(i, column=1).value
    year_of_creation = futurists_sheet.cell(i, column=4).value
    poem_dict = {'author': author, 'year_of_creation': year_of_creation, 'text': text, 'style': 'Футуризм'}
    poetry_dicts.append(poem_dict)
    texts_f.append(text)

futurism_scores = []
missed_poems = []

symbolists_sheet = workbook_with_poems['Символисты']
texts_s = []
for i in range(2,132):
    text = symbolists_sheet.cell(i, column=5).value
    author = symbolists_sheet.cell(i, column=1).value
    year_of_creation = symbolists_sheet.cell(i, column=4).value
    poem_dict = {'author': author, 'year_of_creation': year_of_creation, 'text': text, 'style': 'Символизм'}
    poetry_dicts.append(poem_dict)
    texts_s.append(text)
symbolism_scores = []
missed_poems_s = []

akmeists_sheet = workbook_with_poems['Акмеисты']
texts_a = []
for i in range(2,102):
    text = akmeists_sheet.cell(i, column=5).value
    author = akmeists_sheet.cell(i, column=1).value
    year_of_creation = akmeists_sheet.cell(i, column=4).value
    poem_dict = {'author': author, 'year_of_creation': year_of_creation, 'text': text, 'style': 'Акмеизм'}
    poetry_dicts.append(poem_dict)
    texts_a.append(text)
akmeism_scores = []
missed_poems_a = []

result_dicts = []
counter = 1
#for dictionary in poetry_dicts:
#    print(counter,'из',len(poetry_dicts))
#    result_dict = dictionary
#    text = dictionary['text']
#    futurism_score = many_topics_similarity(text, topics_futurism)
#    result_dict['Футуризм'] = futurism_score
#    symbolism_score = many_topics_similarity(text, topics_symbolism)
#    result_dict['Символизм'] = symbolism_score
#    akmeism_score = many_topics_similarity(text, topics_akmeism)
#    result_dict['Акмеизм'] = akmeism_score
#    scores_dict = {'Футуризм': futurism_score, 'Символизм': symbolism_score, 'Акмеизм': akmeism_score}
#    if get_style_from_scores(scores_dict) == dictionary['style']:
#        result_dict['recognized'] = 1
#    else:
#        result_dict['recognized'] = 0
#    result_dicts.append(result_dict)
#    counter+=1

file_for_results = open('Результаты теста.txt', 'w', encoding='utf-8')

#print(result_dicts, sep='\n', file=file_for_results)

intermedia_6 = '''Привёл себя в упадок. Привёл себя всего.
Стал болен, зелен, гадок. Не то что до того.
Размах, кураж и бодрость утратил наотрез.
Привёл себя в негодность. Низвёл себя с небес.
      
Размяк, померк от пятен. В пепел себя поверг.
Стал неблагоприятен с пятницы по четверг.
Стал сам себе не важен. Сам от себя устал.
Отшелестел плюмажем. Орденом отблистал.

Стал хуже игуаны с шипами на горбу.
Хуже марихуаны и Ленина в гробу.
Против потока двинул. От естества бежал.
Нужное всё отринул. Чуждое всё стяжал.

Блудным явился сыном во изначальный мрак.
Рухнул к таким низинам, ниже каких - никак.
Поднял лицо оттуда. Глянул куда упал.
Понял, что дело худо. Понял - и кончил бал.

Сыграл отбой параду. Велел упадку: стоп.
Украл в аптеке яду. Вылил его в сироп.
Лимон туда же выжал. Спиртом разбавил смесь.
Выпил. Насилу выжил. Но вылечился весь.

Пресёк, пресёк упадок. Выправил статус-кво.
Привёл себя в порядок. Привёл себя в него.
Стал внятен, ладен, годен. Восстановил кураж.
Вернул на место орден. Возобновил плюмаж.

Стал лучше игуаны и даже марабу.
Лучше марихуаны и Ленина в гробу.
Былой престиж удвоил. Пятна на нём замыл.
Нужное всё усвоил. Чуждое всё забыл.

Забыл, как пахнет запах. Забыл, как звук звучит.
Забыл хвататься за бок, если в боку урчит.
Стал точен, прочен, гладок. Не то что перед тем.
Привёл себя в порядок. Извёл себя совсем.

Достиг температуры, близкой к нулю в тени.
Забыл, зачем купюры и где лежат они.
От пагубных повадок навек отвык, отвык.
Увёл себя в осадок. Завёл себя в тупик.

Провёл в себе реформы. Навёл себя на резкость.
Отвёл себе делянку. Обвёл её забором.
Развёл на ней тюльпаны. Довёл до совершенства.
Довёл себя. Таки довёл себя.'''

print('Щербаков',get_style(intermedia_6))

